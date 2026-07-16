"""
Raj Enterprises — Admin Reports Router

Sales reports with daily/weekly/monthly/yearly views.
Export to Excel and PDF.
"""

from fastapi import APIRouter, HTTPException, status, Depends, Query
from fastapi.responses import StreamingResponse
from datetime import datetime, timezone, timedelta
from typing import Optional
from app.database import database
from app.dependencies import require_admin
import io
import logging

logger = logging.getLogger(__name__)
router = APIRouter()


@router.get("/sales")
async def get_sales_report(
    period: str = Query("monthly", regex="^(daily|weekly|monthly|yearly)$"),
    year: int = Query(default=None),
    month: Optional[int] = Query(default=None, ge=1, le=12),
    admin: dict = Depends(require_admin),
):
    """
    Sales report with configurable period.
    Returns aggregated sales data for charts and tables.
    """
    now = datetime.now(timezone.utc)
    if year is None:
        year = now.year

    if period == "daily":
        # Daily sales for a specific month
        if month is None:
            month = now.month
        start = datetime(year, month, 1, tzinfo=timezone.utc)
        if month == 12:
            end = datetime(year + 1, 1, 1, tzinfo=timezone.utc)
        else:
            end = datetime(year, month + 1, 1, tzinfo=timezone.utc)

        pipeline = [
            {"$match": {"created_at": {"$gte": start, "$lt": end}}},
            {"$group": {
                "_id": {"$dayOfMonth": "$created_at"},
                "total_sales": {"$sum": "$amount_total"},
                "total_received": {"$sum": "$amount_received"},
                "order_count": {"$sum": 1},
            }},
            {"$sort": {"_id": 1}},
        ]

    elif period == "weekly":
        # Weekly sales for a specific year
        start = datetime(year, 1, 1, tzinfo=timezone.utc)
        end = datetime(year + 1, 1, 1, tzinfo=timezone.utc)

        pipeline = [
            {"$match": {"created_at": {"$gte": start, "$lt": end}}},
            {"$group": {
                "_id": {"$isoWeek": "$created_at"},
                "total_sales": {"$sum": "$amount_total"},
                "total_received": {"$sum": "$amount_received"},
                "order_count": {"$sum": 1},
            }},
            {"$sort": {"_id": 1}},
        ]

    elif period == "monthly":
        # Monthly sales for a specific year
        start = datetime(year, 1, 1, tzinfo=timezone.utc)
        end = datetime(year + 1, 1, 1, tzinfo=timezone.utc)

        pipeline = [
            {"$match": {"created_at": {"$gte": start, "$lt": end}}},
            {"$group": {
                "_id": {"$month": "$created_at"},
                "total_sales": {"$sum": "$amount_total"},
                "total_received": {"$sum": "$amount_received"},
                "order_count": {"$sum": 1},
            }},
            {"$sort": {"_id": 1}},
        ]

    elif period == "yearly":
        # Yearly sales across all years
        pipeline = [
            {"$group": {
                "_id": {"$year": "$created_at"},
                "total_sales": {"$sum": "$amount_total"},
                "total_received": {"$sum": "$amount_received"},
                "order_count": {"$sum": 1},
            }},
            {"$sort": {"_id": 1}},
        ]

    results = await database.orders.aggregate(pipeline).to_list(length=400)

    # Compute totals
    total_sales = sum(r["total_sales"] for r in results)
    total_received = sum(r["total_received"] for r in results)
    total_orders = sum(r["order_count"] for r in results)

    return {
        "period": period,
        "year": year,
        "month": month,
        "data": [
            {
                "label": r["_id"],
                "total_sales": r["total_sales"],
                "total_received": r["total_received"],
                "total_due": r["total_sales"] - r["total_received"],
                "order_count": r["order_count"],
            }
            for r in results
        ],
        "summary": {
            "total_sales": total_sales,
            "total_received": total_received,
            "total_due": total_sales - total_received,
            "total_orders": total_orders,
        },
    }


@router.get("/sales/export/excel")
async def export_sales_excel(
    period: str = Query("monthly", regex="^(daily|weekly|monthly|yearly)$"),
    year: int = Query(default=None),
    month: Optional[int] = Query(default=None, ge=1, le=12),
    admin: dict = Depends(require_admin),
):
    """Export sales report to Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="openpyxl not installed. Run: pip install openpyxl",
        )

    # Get sales data (reuse the sales endpoint logic)
    now = datetime.now(timezone.utc)
    if year is None:
        year = now.year

    # Simplified: get all orders for the period
    start = datetime(year, 1, 1, tzinfo=timezone.utc)
    end = datetime(year + 1, 1, 1, tzinfo=timezone.utc)
    if month:
        start = datetime(year, month, 1, tzinfo=timezone.utc)
        end = datetime(year, month + 1, 1, tzinfo=timezone.utc) if month < 12 else datetime(year + 1, 1, 1, tzinfo=timezone.utc)

    orders = await database.orders.find(
        {"created_at": {"$gte": start, "$lt": end}}
    ).sort("created_at", -1).to_list(length=5000)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")

    headers = ["Order #", "Date", "Customer", "Items", "Total", "Received", "Due", "Status", "Payment Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for row, order in enumerate(orders, 2):
        user = await database.users.find_one({"_id": order["user_id"]})
        ws.cell(row=row, column=1, value=order["order_number"])
        ws.cell(row=row, column=2, value=order["created_at"].strftime("%Y-%m-%d"))
        ws.cell(row=row, column=3, value=user["name"] if user else "Unknown")
        ws.cell(row=row, column=4, value=len(order.get("items", [])))
        ws.cell(row=row, column=5, value=order["amount_total"])
        ws.cell(row=row, column=6, value=order.get("amount_received", 0))
        ws.cell(row=row, column=7, value=max(0, order["amount_total"] - order.get("amount_received", 0)))
        ws.cell(row=row, column=8, value=order["order_status"])
        ws.cell(row=row, column=9, value=order["payment_status"])

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 30)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"sales_report_{year}{'_' + str(month) if month else ''}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/sales/export/pdf")
async def export_sales_pdf(
    period: str = Query("monthly", regex="^(daily|weekly|monthly|yearly)$"),
    year: int = Query(default=None),
    month: Optional[int] = Query(default=None, ge=1, le=12),
    admin: dict = Depends(require_admin),
):
    """Export sales report to PDF."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="reportlab not installed. Run: pip install reportlab",
        )

    now = datetime.now(timezone.utc)
    if year is None:
        year = now.year

    start = datetime(year, 1, 1, tzinfo=timezone.utc)
    end = datetime(year + 1, 1, 1, tzinfo=timezone.utc)
    if month:
        start = datetime(year, month, 1, tzinfo=timezone.utc)
        end = datetime(year, month + 1, 1, tzinfo=timezone.utc) if month < 12 else datetime(year + 1, 1, 1, tzinfo=timezone.utc)

    orders = await database.orders.find(
        {"created_at": {"$gte": start, "$lt": end}}
    ).sort("created_at", -1).to_list(length=5000)

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    elements = []

    # Title
    title = f"{settings.company_name} — Sales Report"
    elements.append(Paragraph(title, styles["Title"]))
    elements.append(Paragraph(f"Period: {year}" + (f"-{month:02d}" if month else ""), styles["Normal"]))
    elements.append(Spacer(1, 20))

    # Table
    data = [["Order #", "Date", "Total", "Received", "Due", "Status", "Payment"]]
    for order in orders:
        data.append([
            order["order_number"],
            order["created_at"].strftime("%Y-%m-%d"),
            f"₹{order['amount_total']:.2f}",
            f"₹{order.get('amount_received', 0):.2f}",
            f"₹{max(0, order['amount_total'] - order.get('amount_received', 0)):.2f}",
            order["order_status"],
            order["payment_status"],
        ])

    table = Table(data)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E4057")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
        ("GRID", (0, 0), (-1, -1), 1, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
    ]))
    elements.append(table)

    doc.build(elements)
    output.seek(0)

    filename = f"sales_report_{year}{'_' + str(month) if month else ''}.pdf"

    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


from bson import ObjectId

@router.get("/dashboard")
async def get_dashboard_summary(admin: dict = Depends(require_admin)):
    """Retrieve overall stats dashboard metrics."""
    sales_pipeline = [
        {"$match": {"order_status": {"$ne": "cancelled"}}},
        {"$group": {"_id": None, "total": {"$sum": "$amount_total"}}}
    ]
    sales_res = await database.orders.aggregate(sales_pipeline).to_list(length=1)
    total_sales = sales_res[0]["total"] if sales_res else 0.0

    dues_pipeline = [
        {"$match": {"order_status": {"$ne": "cancelled"}}},
        {"$group": {
            "_id": None,
            "total_received": {"$sum": "$amount_received"},
            "total_amount": {"$sum": "$amount_total"}
        }}
    ]
    dues_res = await database.orders.aggregate(dues_pipeline).to_list(length=1)
    total_dues = max(0.0, dues_res[0]["total_amount"] - dues_res[0]["total_received"]) if dues_res else 0.0

    total_orders = await database.orders.count_documents({})
    active_orders = await database.orders.count_documents({
        "order_status": {"$in": ["placed", "confirmed", "packed", "dispatched"]}
    })

    low_stock_count = 0
    async for p in database.products.find({"status": {"$ne": "inactive"}}):
        if p.get("stock_count", 0) <= p.get("low_stock_threshold", 5):
            low_stock_count += 1

    return {
        "total_sales": total_sales,
        "total_dues": total_dues,
        "total_orders": total_orders,
        "active_orders": active_orders,
        "low_stock_count": low_stock_count,
    }


@router.get("/sales-trends")
async def get_sales_trends(admin: dict = Depends(require_admin)):
    """Retrieve monthly sales trend series for the current year."""
    now = datetime.now(timezone.utc)
    year_start = datetime(now.year, 1, 1, tzinfo=timezone.utc)

    pipeline = [
        {"$match": {
            "created_at": {"$gte": year_start},
            "order_status": {"$ne": "cancelled"}
        }},
        {"$group": {
            "_id": {"month": {"$month": "$created_at"}},
            "sales": {"$sum": "$amount_total"}
        }},
        {"$sort": {"_id.month": 1}}
    ]

    db_res = await database.orders.aggregate(pipeline).to_list(length=12)
    
    months_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    series = [{"month": name, "sales": 0.0} for name in months_names]

    for item in db_res:
        m_idx = item["_id"]["month"] - 1
        if 0 <= m_idx < 12:
            series[m_idx]["sales"] = item["sales"]

    return {"series": series}


@router.get("/top-products")
async def get_top_products(admin: dict = Depends(require_admin)):
    """Retrieve top-selling products by quantity and revenue."""
    pipeline = [
        {"$match": {"order_status": {"$ne": "cancelled"}}},
        {"$unwind": "$items"},
        {"$group": {
            "_id": "$items.product_id",
            "title": {"$first": "$items.title_snapshot"},
            "units_sold": {"$sum": "$items.quantity"},
            "revenue": {"$sum": "$items.subtotal"}
        }},
        {"$sort": {"units_sold": -1}},
        {"$limit": 10}
    ]

    db_res = await database.orders.aggregate(pipeline).to_list(length=10)
    
    formatted = []
    for item in db_res:
        sku = "RE-PAINT"
        product = await database.products.find_one({"_id": ObjectId(item["_id"])})
        if product:
            sku = product.get("sku", sku)

        formatted.append({
            "product_id": str(item["_id"]),
            "sku": sku,
            "title": item["title"],
            "units_sold": item["units_sold"],
            "revenue": item["revenue"]
        })

    return {"products": formatted}
